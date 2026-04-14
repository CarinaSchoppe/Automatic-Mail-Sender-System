from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from mail_sender import cli
from mail_sender.recipients import Recipient


def write_recipient(path: Path, company: str, email: str) -> None:
    path.write_text(f"company,mail\n{company},{email}\n", encoding="utf-8")


def test_cli_auto_processes_each_input_folder_and_logs_dry_run(project: Path, capsys) -> None:
    write_recipient(project / "input/PhD/phd.csv", "PhD Co", "phd@example.com")
    write_recipient(project / "input/Freelance_German/de.txt", "DE Co", "de@example.com")
    write_recipient(project / "input/Freelance_English/en.csv", "EN Co", "en@example.com")

    result = cli.main([
        "--mode",
        "Auto",
        "--base-dir",
        str(project),
        "--allow-empty-attachments",
        "--log-dry-run",
        "--verbose",
    ])

    assert result == 0
    output = capsys.readouterr().out
    assert "Mode: PhD" in output
    assert "Mode: Freelance German" in output
    assert "Mode: Freelance English" in output
    assert load_workbook(project / "output/send_phd.xlsx").active.max_row == 2
    assert load_workbook(project / "output/send_freelance.xlsx").active.max_row == 3


def test_cli_skips_logged_and_duplicate_addresses(project: Path, capsys) -> None:
    write_recipient(project / "input/PhD/one.csv", "One", "one@example.com")
    (project / "input/PhD/two.csv").write_text("company,mail\nTwo,one@example.com\nLogged,logged@example.com\n", encoding="utf-8")
    workbook = Workbook()
    workbook.active.append(["Unternehmen", "mail", "sent_at"])
    workbook.active.append(["Logged", "mailto:logged@example.com", "2026-04-14T10:00+10:00"])
    workbook.save(project / "output/send_phd.xlsx")

    result = cli.main(["--mode", "PhD", "--base-dir", str(project), "--allow-empty-attachments"])

    assert result == 0
    output = capsys.readouterr().out
    assert "duplicate skipped" in output
    assert "already in send_phd.xlsx" in output
    assert "[DRY_RUN] one@example.com" in output


def test_cli_returns_zero_when_all_recipients_are_logged(project: Path, capsys) -> None:
    write_recipient(project / "input/PhD/one.csv", "One", "one@example.com")
    workbook = Workbook()
    workbook.active.append(["Unternehmen", "mail", "sent_at"])
    workbook.active.append(["One", "one@example.com", "2026-04-14T10:00+10:00"])
    workbook.save(project / "output/send_phd.xlsx")

    result = cli.main(["--mode", "PhD", "--base-dir", str(project), "--allow-empty-attachments", "--verbose"])

    assert result == 0
    output = capsys.readouterr().out
    assert "Loaded 1 existing email address" in output
    assert "Nothing to process." in output


def test_cli_returns_zero_when_auto_has_no_files(project: Path, capsys) -> None:
    result = cli.main(["--mode", "Auto", "--base-dir", str(project)])

    assert result == 0
    assert "No input files found" in capsys.readouterr().out


def test_cli_specific_mode_reports_missing_input_files(project: Path, capsys) -> None:
    result = cli.main(["--mode", "PhD", "--base-dir", str(project), "--allow-empty-attachments", "--verbose"])

    assert result == 1
    output = capsys.readouterr().out
    assert "No recipient files found." in output
    assert "No .csv or .txt recipient files found" in output


def test_cli_errors_for_missing_attachments(project: Path, capsys) -> None:
    write_recipient(project / "input/Freelance_English/en.csv", "EN Co", "en@example.com")

    result = cli.main(["--mode", "Freelance_English", "--base-dir", str(project)])

    assert result == 1
    assert "No attachments found" in capsys.readouterr().out


def test_cli_resend_existing_bypasses_output_log(project: Path, capsys) -> None:
    write_recipient(project / "input/PhD/phd.csv", "PhD Co", "phd@example.com")
    result = cli.main([
        "--mode",
        "PhD",
        "--base-dir",
        str(project),
        "--allow-empty-attachments",
        "--resend-existing",
        "--verbose",
    ])

    assert result == 0
    assert "Existing Excel log addresses will be ignored" in capsys.readouterr().out


def test_cli_send_path_uses_mailer(monkeypatch, project: Path) -> None:
    write_recipient(project / "input/PhD/phd.csv", "PhD Co", "phd@example.com")
    (project / "attachments/PhD/file.txt").write_text("attachment", encoding="utf-8")
    sent = []

    class FakeMailer:
        def __init__(self, config) -> None:
            self.config = config

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback) -> None:
            return None

        def send(self, recipient, subject, text_body, html_body, attachments, inline_images) -> None:
            sent.append((recipient.email, subject, len(attachments), len(inline_images)))

    monkeypatch.setenv("SMTP_PASSWORD", "secret")
    monkeypatch.setattr("mail_sender.cli.SmtpMailer", FakeMailer)

    result = cli.main(["--mode", "PhD", "--base-dir", str(project), "--send"])

    assert result == 0
    assert sent == [("phd@example.com", "PhD PhD Co", 1, 1)]
    sheet = load_workbook(project / "output/send_phd.xlsx").active
    assert sheet.cell(2, 2).value == "phd@example.com"


def test_cli_can_disable_sent_excel_logging(monkeypatch, project: Path) -> None:
    write_recipient(project / "input/PhD/phd.csv", "PhD Co", "phd@example.com")
    (project / "attachments/PhD/file.txt").write_text("attachment", encoding="utf-8")

    class FakeMailer:
        def __init__(self, config) -> None:
            self.config = config

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback) -> None:
            return None

        def send(self, recipient, subject, text_body, html_body, attachments, inline_images) -> None:
            return None

    monkeypatch.setenv("SMTP_PASSWORD", "secret")
    monkeypatch.setattr("mail_sender.cli.SmtpMailer", FakeMailer)

    result = cli.main(["--mode", "PhD", "--base-dir", str(project), "--send", "--no-write-sent-log"])

    assert result == 0
    assert not (project / "output/send_phd.xlsx").exists()


def test_cli_returns_error_when_processing_recipient_fails(monkeypatch, project: Path, capsys) -> None:
    write_recipient(project / "input/PhD/phd.csv", "PhD Co", "phd@example.com")

    def broken_render(*args, **kwargs):
        raise ValueError("boom")

    monkeypatch.setattr("mail_sender.cli.render_mail", broken_render)

    result = cli.main(["--mode", "PhD", "--base-dir", str(project), "--allow-empty-attachments"])

    assert result == 1
    assert "[ERROR] phd@example.com | boom" in capsys.readouterr().out


def test_process_recipients_requires_mailer_when_not_dry_run(project: Path) -> None:
    errors = cli._process_recipients(
        mailer=None,
        template_path=project / "templates/phd.txt",
        signature_path=project / "templates/signature.txt",
        log_path=project / "output/send_phd.xlsx",
        recipients=[],
        attachments=[],
        subject_override=None,
        signature_image_path=project / "templates/signature-logo.png",
        signature_image_width=180,
        dry_run=False,
        log_dry_run=False,
        write_sent_log=False,
        verbose=False,
    )
    assert errors == 0

    errors = cli._process_recipients(
        mailer=None,
        template_path=project / "templates/phd.txt",
        signature_path=project / "templates/signature.txt",
        log_path=project / "output/send_phd.xlsx",
        recipients=[Recipient(email="a@example.com", company="A")],
        attachments=[],
        subject_override=None,
        signature_image_path=project / "templates/signature-logo.png",
        signature_image_width=180,
        dry_run=False,
        log_dry_run=False,
        write_sent_log=False,
        verbose=False,
    )
    assert errors == 1
