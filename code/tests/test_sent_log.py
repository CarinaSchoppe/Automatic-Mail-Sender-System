from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from mail_sender.recipients import Recipient
from mail_sender.sent_log import append_invalid_email, append_log, read_invalid_emails, read_known_output_emails, read_logged_emails


def test_append_log_creates_three_column_output(tmp_path: Path) -> None:
    log_path = tmp_path / "output/send_phd.xlsx"

    append_log(log_path, Recipient(email="person@example.com", company="ACME"))

    sheet = load_workbook(log_path).active
    assert sheet.max_column == 3
    assert [sheet.cell(1, column).value for column in range(1, 4)] == ["company", "mail", "sent_at"]
    assert [sheet.cell(2, column).value for column in range(1, 3)] == ["ACME", "person@example.com"]
    assert "+10:00" in sheet.cell(2, 3).value
    assert read_logged_emails(log_path) == {"person@example.com"}


def test_read_logged_emails_normalizes_mailto_and_handles_missing_headers(tmp_path: Path) -> None:
    assert read_logged_emails(tmp_path / "missing.xlsx") == set()

    log_path = tmp_path / "send.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["company", "mail"])
    sheet.append(["ACME", "mailto:person@example.com"])
    sheet.append(["Blank", ""])
    workbook.save(log_path)

    assert read_logged_emails(log_path) == {"person@example.com"}

    no_mail = tmp_path / "no-mail.xlsx"
    workbook = Workbook()
    workbook.active.append(["company"])
    workbook.save(no_mail)
    assert read_logged_emails(no_mail) == set()


def test_append_log_truncates_old_extra_columns(tmp_path: Path) -> None:
    log_path = tmp_path / "send.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["company", "mail", "sent_at_utc", "mode", "status"])
    sheet.append(["Old", "old@example.com", "old", "PhD", "SENT"])
    workbook.save(log_path)

    append_log(log_path, Recipient(email="new@example.com", company="New"))

    sheet = load_workbook(log_path).active
    assert sheet.max_column == 3
    assert [sheet.cell(1, column).value for column in range(1, 4)] == ["company", "mail", "sent_at"]


def test_invalid_email_log_and_known_output_email_scan(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    sent_path = output_dir / "send_phd.xlsx"
    invalid_path = output_dir / "invalid_mails.xlsx"

    append_log(sent_path, Recipient(email="sent@example.com", company="Sent"))
    append_invalid_email(invalid_path, Recipient(email="bad@example.invalid", company="Bad"), "domain has no MX or A record")

    invalid_sheet = load_workbook(invalid_path).active
    assert invalid_sheet.title == "Invalid"
    assert [invalid_sheet.cell(1, column).value for column in range(1, 5)] == [
        "company",
        "mail",
        "invalid_reason",
        "detected_at",
    ]
    assert read_invalid_emails(invalid_path) == {"bad@example.invalid"}
    assert read_known_output_emails(output_dir) == {"sent@example.com"}


def test_known_output_email_scan_handles_missing_output_dir(tmp_path: Path) -> None:
    assert read_known_output_emails(tmp_path / "missing-output") == set()
