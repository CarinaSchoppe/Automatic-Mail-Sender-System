from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from mail_sender.recipients import EMAIL_KEYS
from mail_sender.recipients import Recipient
from mail_sender.recipients import normalize_email
from mail_sender.recipients import normalize_key

HEADERS = [
    "company",
    "mail",
    "sent_at",
]

INVALID_HEADERS = [
    "company",
    "mail",
    "invalid_reason",
    "detected_at",
]


def append_log(
        log_path: Path,
        recipient: Recipient,
) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    workbook, sheet = _open_or_create_workbook(log_path)
    _ensure_headers(sheet)
    sheet.append(
        [
            recipient.company,
            recipient.email,
            datetime.now(ZoneInfo("Australia/Brisbane")).isoformat(timespec="minutes"),
        ]
    )
    workbook.save(log_path)


def append_invalid_email(log_path: Path, recipient: Recipient, reason: str) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    workbook, sheet = _open_or_create_workbook(log_path, INVALID_HEADERS, "Invalid")
    _ensure_headers(sheet, INVALID_HEADERS)
    sheet.append(
        [
            recipient.company,
            recipient.email,
            reason,
            datetime.now(ZoneInfo("Australia/Brisbane")).isoformat(timespec="minutes"),
        ]
    )
    workbook.save(log_path)


def read_logged_emails(log_path: Path) -> set[str]:
    if not log_path.exists():
        return set()

    workbook = load_workbook(log_path)
    sheet = workbook.active
    header = [str(cell.value or "") for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    email_index = _find_header_index(header, EMAIL_KEYS)
    if email_index is None:
        return set()

    emails: set[str] = set()
    for row in sheet.iter_rows(min_row=2):
        email = normalize_email(str(row[email_index - 1].value or "")).lower()
        if email:
            emails.add(email)
    return emails


def read_invalid_emails(log_path: Path) -> set[str]:
    return read_logged_emails(log_path)


def read_known_output_emails(output_dir: Path) -> set[str]:
    if not output_dir.exists() or not output_dir.is_dir():
        return set()

    emails: set[str] = set()
    for path in output_dir.glob("*.xlsx"):
        if path.name.lower() == "invalid_mails.xlsx":
            continue
        emails.update(read_logged_emails(path))
    return emails


def _open_or_create_workbook(
        log_path: Path,
        headers: list[str] | None = None,
        title: str = "Sent",
) -> tuple[Workbook, Worksheet]:
    headers = headers or HEADERS
    if log_path.exists():
        workbook = load_workbook(log_path)
        return workbook, workbook.active

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    return workbook, sheet


def _ensure_headers(sheet: Worksheet, headers: list[str] | None = None) -> None:
    headers = headers or HEADERS
    if sheet.max_column > len(headers):
        sheet.delete_cols(len(headers) + 1, sheet.max_column - len(headers))

    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=header)


def _find_header_index(header: list[str], allowed_keys: set[str]) -> int | None:
    for index, value in enumerate(header, start=1):
        if normalize_key(value) in allowed_keys:
            return index
    return None
