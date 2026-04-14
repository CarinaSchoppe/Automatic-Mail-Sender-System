from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from mail_sender.recipients import Recipient
from mail_sender.recipients import EMAIL_KEYS
from mail_sender.recipients import normalize_key


HEADERS = [
    "Unternehmen",
    "mail",
    "sent_at_utc",
    "mode",
    "status",
    "subject",
    "attachments",
    "error",
]


def append_log(
    log_path: Path,
    mode: str,
    status: str,
    recipient: Recipient,
    subject: str,
    attachments: list[Path],
    error: str = "",
) -> None:
    workbook, sheet = _open_or_create_workbook(log_path)
    _ensure_headers(sheet)
    sheet.append(
        [
            recipient.company,
            recipient.email,
            datetime.now(timezone.utc).isoformat(timespec="seconds"),
            mode,
            status,
            subject,
            "; ".join(path.name for path in attachments),
            error,
        ]
    )
    workbook.save(log_path)


def read_logged_emails(log_path: Path) -> set[str]:
    if not log_path.exists():
        return set()

    workbook = load_workbook(log_path)
    sheet = workbook.active
    header = [str(cell.value or "") for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    email_index = _find_header_index(header, EMAIL_KEYS)
    if email_index is None:
        return set()

    emails: set[str] = set()
    for row in sheet.iter_rows(min_row=2):
        email = str(row[email_index - 1].value or "").strip().lower()
        if email:
            emails.add(email)
    return emails


def _open_or_create_workbook(log_path: Path) -> tuple[Workbook, Worksheet]:
    if log_path.exists():
        workbook = load_workbook(log_path)
        return workbook, workbook.active

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sent"
    sheet.append(HEADERS)
    return workbook, sheet


def _ensure_headers(sheet: Worksheet) -> None:
    existing = [str(cell.value or "") for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized_existing = {normalize_key(value) for value in existing}

    for header in HEADERS:
        if normalize_key(header) not in normalized_existing:
            sheet.cell(row=1, column=sheet.max_column + 1, value=header)
            normalized_existing.add(normalize_key(header))


def _find_header_index(header: list[str], allowed_keys: set[str]) -> int | None:
    for index, value in enumerate(header, start=1):
        if normalize_key(value) in allowed_keys:
            return index
    return None
